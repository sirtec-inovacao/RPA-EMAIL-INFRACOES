import win32com.client as win32
import pandas as pd
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import pyautogui
from PIL import ImageGrab
from time import sleep
import glob


# Variaveis
file_path = r'C:\Users\luana\Documents\GitHub\robo-infracoes\src\ponto_mais\analysis\Analise_Excel\Analise.xlsx'
pasta_analise = "C:/Users/luana/Documents/GitHub/robo-infracoes/src/ponto_mais/analysis/Analise_Excel/Analise.xlsx"

gerentes_mensal = {
    "FRS": "Mensal - Ocorrencias de ponto - FRS.xlsx",
    "BAR": "Mensal - Ocorrencias de ponto - BAR.xlsx",
    "BJL": "Mensal - Ocorrencias de ponto - BJL.xlsx",
    "VTC": "Mensal - Ocorrencias de ponto - SUD.xlsx",
    "CE": "Mensal - Ocorrencias de ponto - CE.xlsx",
    "RS": "Mensal - Ocorrencias de ponto - RS.xlsx",
    "PEL": "Mensal - Ocorrencias de ponto - PEL.xlsx",
}    

gerentes = {
    "FRS": "Ocorrencias de ponto - FRS.xlsx",
    "BAR": "Ocorrencias de ponto - BAR.xlsx",
    "BJL": "Ocorrencias de ponto - BJL.xlsx",
    "VTC": "Ocorrencias de ponto - SUD.xlsx",
    "CE": "Ocorrencias de ponto - CE.xlsx",
    "RS": "Ocorrencias de ponto - RS.xlsx",
    "PEL": "Ocorrencias de ponto - PEL.xlsx",
}

pasta_saida = {
    "FRS": "C:/Users/luana/Documents/GitHub/robo-infracoes/src/ponto_mais/analysis/FRS",
    "BAR": "C:/Users/luana/Documents/GitHub/robo-infracoes/src/ponto_mais/analysis/BAR",
    "BJL": "C:/Users/luana/Documents/GitHub/robo-infracoes/src/ponto_mais/analysis/BJL",
    "VTC": "C:/Users/luana/Documents/GitHub/robo-infracoes/src/ponto_mais/analysis/VTC",
    "CE": "C:/Users/luana/Documents/GitHub/robo-infracoes/src/ponto_mais/analysis/CE",
    "RS": "C:/Users/luana/Documents/GitHub/robo-infracoes/src/ponto_mais/analysis/RS",
    "PEL": "C:/Users/luana/Documents/GitHub/robo-infracoes/src/ponto_mais/analysis/PEL",
}

def atualizar_planilha_analise(file_path):
    
    # Abrindo o Excel
    print("Abrindo arquivo de Analise")
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = True

    # Abrindo o arquivo Excel
    workbook = excel.Workbooks.Open(file_path)
    sleep(15)
    print("Atualizando dados")
    
    # Executando o comando "Atualizar Tudo"
    excel.ActiveWorkbook.RefreshAll()

    # Esperando 3 minutos para as atualizações terminarem
    sleep(180)

    # Salvando e fechando o arquivo
    print("Dados atualizados, salvando e fechando o arquivo")
    workbook.Save()
    sleep(2)
    workbook.Close()

    # Fechando o Excel
    excel.Quit()
    

def processar_gerente_finalizada(gerente, pasta_analise, pasta_saida):
    # Caminho do arquivo de saída
    send_file = pasta_saida[gerente] + f"/Ocorrencias de ponto - {gerente}.xlsx"

    # Carregar o arquivo Excel
    wb_analise = pd.read_excel(pasta_analise, sheet_name="Finalizada", engine='openpyxl', header=1)
    sleep(15)
    
    # Filtrar os dados
    df_filtrado = wb_analise[(wb_analise['GERENTE'].isin([gerente])) & (wb_analise['VALIDA'].isin(["SIM"]))]
    
    # Ordenar os dados pela coluna "PONTUAÇÃO" do maior ao menor
    df_filtrado = df_filtrado.sort_values(by="PONTUACAO", ascending=False)

    # Criar ou abrir o arquivo de saída
    wb_send = openpyxl.load_workbook(send_file) if os.path.exists(send_file) else openpyxl.Workbook()
    ws_send = wb_send.active

    # Limpar os dados existentes na send_file
    ws_send.delete_rows(3, ws_send.max_row)

    # Preencher a aba com os novos dados
    tabela = pd.DataFrame(df_filtrado.iloc[:, :12].values, columns=df_filtrado.columns[:12])
    for row in dataframe_to_rows(tabela, index=False, header=False):
        ws_send.append(row)

    # Aplicar formatação de data
    date_format = "DD/MM/YYYY"
    for row in ws_send.iter_rows(min_row=3, max_row=ws_send.max_row):
        for cell in row:
            if isinstance(cell.value, pd.Timestamp):
                cell.number_format = date_format

    # Aplicar formatação de centralização em todas as células a partir da linha 3
    for row in ws_send.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # Salvar as alterações na send_file
    wb_send.save(send_file)

    print(f"Dados filtrados e colados na {send_file} com sucesso.")

    # Abrir a planilha Excel manualmente usando o PyAutoGUI
    pyautogui.hotkey('winleft', 'r')  # Pressionar Win + R para abrir a caixa de execução
    sleep(2)
    pyautogui.typewrite(f'excel.exe "{send_file}"')
    sleep(2)
    pyautogui.press('enter')  # Pressionar Enter para abrir o arquivo

    # Aguardar um tempo para dar tempo de abrir a planilha
    sleep(10)

    # Dimensões do print
    left = 25
    top = 280
    right = 1340
    bottom = 465

# Salvando arquivo .png nas pastas de saída
    screenshot = ImageGrab.grab(bbox=(left, top, right, bottom))
    screenshot.save(f'{pasta_saida[gerente]}/Envio de ocorrencia dos pontos {gerente}.png')
    sleep(10)
    pyautogui.hotkey('alt', 'f4')
    
    
def processar_gerente_mensal(gerente, pasta_analise, pasta_saida):
    # Caminho do arquivo de saída
    send_file = pasta_saida[gerente] + f"/Mensal - Ocorrencias de ponto - {gerente}.xlsx"

    # Carregar o arquivo Excel
    wb_analise = pd.read_excel(pasta_analise, sheet_name="Analise_Completa", engine='openpyxl', header=1)
    sleep(15)
    
    # Filtrar os dados
    df_filtrado = wb_analise[(wb_analise['GERENTE'].isin([gerente])) & (wb_analise['VALIDA'].isin(["SIM"]))]
    
    # Ordenar os dados pela coluna "PONTUAÇÃO" do maior ao menor
    df_filtrado = df_filtrado.sort_values(by="PONTUACAO", ascending=False)

    # Criar ou abrir o arquivo de saída
    wb_send = openpyxl.load_workbook(send_file) if os.path.exists(send_file) else openpyxl.Workbook()
    ws_send = wb_send.active

    # Limpar os dados existentes na send_file
    ws_send.delete_rows(3, ws_send.max_row)

    # Preencher a aba com os novos dados
    tabela = pd.DataFrame(df_filtrado.iloc[:, :12].values, columns=df_filtrado.columns[:12])
    for row in dataframe_to_rows(tabela, index=False, header=False):
        ws_send.append(row)

    # Aplicar formatação de data
    date_format = "DD/MM/YYYY"
    for row in ws_send.iter_rows(min_row=3, max_row=ws_send.max_row):
        for cell in row:
            if isinstance(cell.value, pd.Timestamp):
                cell.number_format = date_format

    # Aplicar formatação de centralização em todas as células a partir da linha 3
    for row in ws_send.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # Salvar as alterações na send_file
    wb_send.save(send_file)

    print(f"Dados filtrados e colados na {send_file} com sucesso.")

    # Abrir a planilha Excel manualmente usando o PyAutoGUI
    pyautogui.hotkey('winleft', 'r')  # Pressionar Win + R para abrir a caixa de execução
    sleep(2)
    pyautogui.typewrite(f'excel.exe "{send_file}"')
    sleep(2)
    pyautogui.press('enter')  # Pressionar Enter para abrir o arquivo

    # Aguardar um tempo para dar tempo de abrir a planilha
    sleep(10)

    # Dimensões do print
    left = 25
    top = 280
    right = 1340
    bottom = 465

# Salvando arquivo .png nas pastas de saída
    screenshot = ImageGrab.grab(bbox=(left, top, right, bottom))
    screenshot.save(f'{pasta_saida[gerente]}/Mensal - Envio de ocorrencia dos pontos {gerente}.png')
    sleep(10)
    pyautogui.hotkey('alt', 'f4')
    

def analise_main():
        # Iterar sobre os gerentes
    for gerente in gerentes.keys():
        processar_gerente_finalizada(gerente, pasta_analise, pasta_saida)
        
    # Iterar sobre os gerentes
    for gerente in gerentes_mensal.keys():
        processar_gerente_mensal(gerente, pasta_analise, pasta_saida)

    atualizar_planilha_analise(file_path)